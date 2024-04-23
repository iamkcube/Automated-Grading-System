from openpyxl import load_workbook
from PIL import Image, ImageFont, ImageDraw
import pandas as pd
from xhtml2pdf import pisa

import sys
import os
parentdir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, fr'{parentdir}/recognition')
from prediction import get_prediction
from barcode_reader import barcode_reader


'''Global Variables'''
# FONT_FILE = ImageFont.truetype(r"font/Silentha OT.ttf", 150)
# FONT_COLOR = "#FFFFFF"

# sys.path.insert(0, fr'{parentdir}/database')
# print(sys.path)
print(os.getcwd())
marks_excel_sheet_filename = r"C:\Users\kalinga\Downloads\Mine\All Codes\AutomaticGradingSystem\database\Marks.xlsx"
students_excel_sheet_filename = r"C:\Users\kalinga\Downloads\Mine\All Codes\AutomaticGradingSystem\database\Students.xlsx"
marks_excel_sheet = load_workbook(filename=marks_excel_sheet_filename)
students_excel_sheet = load_workbook(filename=students_excel_sheet_filename)


def update_marks_sheet(student_id, subject_code, internal_marks = None,endsem_marks=None, total_marks=None, grades=None):
    sheet = marks_excel_sheet.active
    print(type(student_id), subject_code, endsem_marks)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
        print("row= ", type(row[0].value), row[1].value)
        if row[0].value == int(student_id) and row[1].value == subject_code:
            print("2:", row[2].value)
            print("3:", row[3].value)
            if internal_marks:
                row[3].value = internal_marks
                row[5].value = internal_marks + row[4].value
                row[6].value = getGrade(internal_marks + row[4].value)
            if endsem_marks:
                row[4].value = endsem_marks
                row[5].value = row[3].value + endsem_marks
                row[6].value = getGrade(row[3].value + endsem_marks)
            break

    marks_excel_sheet.save(filename=marks_excel_sheet_filename)


def update_students_sheet(student_id, sem, value):
    sheet = students_excel_sheet.active

    row_num = 0

    if sem == 1:
        row_num = 4
    elif sem == 2:
        row_num = 5
    elif sem == 3:
        row_num = 6
    elif sem == 4:
        row_num = 7
    elif sem == 5:
        row_num = 8
    elif sem == 6:
        row_num = 9
    elif sem == 7:
        row_num = 10
    elif sem == 8:
        row_num = 11
    else:
        row_num = 12

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13):
        if row[0].value == student_id:
            row[row_num].value = value
            break

    students_excel_sheet.save(filename=students_excel_sheet_filename)

def getGrade(marks):
    if marks >= 90:
        return 'O'
    elif 80 <= marks < 90:
        return 'E'
    elif 70 <= marks < 80:
        return 'A'
    elif 60 <= marks < 70:
        return 'B'
    elif 50 <= marks < 60:
        return 'C'
    elif 40 <= marks < 50:
        return 'D'
    else:
        return 'F'

def getCredits(sem=0):
    if sem == 1:
        return 20.5
    elif sem == 2:
        return 20.5
    elif sem == 3:
        return 23
    elif sem == 4:
        return 20.5
    elif sem == 5:
        return 22.5
    elif sem == 6:
        return 20
    elif sem == 7:
        return 19
    elif sem == 8:
        return 17


def updateCGPA(student_id):
    sheet = students_excel_sheet.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13, values_only=True):
        if row[0] == student_id:
            total_cgpa = 0
            total_credits = 0
            # sum([(row[i]*getCredits(i-3))/getCredits() for i in range(4, 12) if row[i] is not None])

            for i in range(4, 12):
                if row[i] is None:
                    break
                credits = getCredits(i-3)
                total_cgpa += row[i] * credits
                total_credits += credits

            total_cgpa = total_cgpa/total_credits if total_credits != 0 else 0

            print(total_cgpa)

    update_students_sheet(student_id=student_id, sem=0,
                          value=f"{total_cgpa:.2f}")


def make_certificate(student_id):
    students_sheet = students_excel_sheet.active
    marks_sheet = marks_excel_sheet.active

    marks_dict = {}

    students_df = pd.read_excel(students_excel_sheet_filename)
    marks_df = pd.read_excel(marks_excel_sheet_filename)

    filtered_data = students_df[students_df['Student ID'] == int(student_id)]
    if not filtered_data.empty:
        name = filtered_data.iloc[0]['Student Name']
        cgpa = filtered_data.iloc[0]['CGPA']

    student_101_data = marks_df[marks_df['Student ID'] == int(student_id)]
    grades_dict = student_101_data.set_index(
        'Subject Code')['Grades'].to_dict()

    marks_dict["student_id"] = int(student_id)
    marks_dict["name"] = name
    marks_dict["cgpa"] = cgpa
    marks_dict["grades"] = grades_dict

    convert_html_to_pdf(marks_dict=marks_dict,
                        pdf_path=fr"C:\Users\kalinga\Downloads\Mine\All Codes\AutomaticGradingSystem\backend\certificates/Certificate-{student_id}.pdf")


def convert_html_to_pdf(marks_dict, pdf_path):
    with open(pdf_path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(
            make_html_string(marks_dict), dest=pdf_file)

    return not pisa_status.err


def make_html_string(marks_dict):
    subjects_string = ""

    for index, (subject, grade) in enumerate(marks_dict["grades"].items(), 1):
        subjects_string += f'''
                    <tr>
                        <td>{index}</td>
                        <td>{subject.upper()}</td>
                        <td>{grade}</td>
                    </tr>'''

    html_string = f'''
    <!DOCTYPE html>
    <html lang="en">
        <style>
            * {{
                margin: 0;
                padding: 0;
            }}
            body {{
                margin: auto;
                margin-block: 2rem;
            }}
            .header-container {{
                margin: 0;
                padding: 0;
            }}
            .header-img{{
                margin: auto;
            }}
            .flag {{
                display: grid;
                background: #007a16;
                height: 41px;
                overflow: hidden;
                border-bottom: 2px solid #70d06e;
            }}
            .flag_left h2 {{
                color: #70d06e;
                padding: 7px 9px;
                font-size: 20px;
            }}
            .green_background {{
                background: #007a16;
                padding: 4px 9px;
                font-size: 22px;
                color: #fff;
            }}
            .official {{
                text-align: right;
                background: #70d06e;
                padding: 5px;
                text-decoration: none;
                color: #fff;
                font-size: 14px;
            }}
            .fast_heading {{
                text-align: center;
                padding: 20px 0;
                padding-bottom: 2px;
                border-bottom: 2px solid #eee;
                font-size: 2rem;
                margin-top: 12px;
            }}

            .student_detials table {{
                background: #eee;
                border-collapse: collapse;
                width: 100%;
                margin-top: 40px;
            }}
            .student_detials table tr td {{
                border: 1px solid #fff;
                font-size: 14px;
                padding: 3px 3px;
            }}
            .second_heading {{
                text-align: center;
                margin-top: 20px;
            }}
            .second_heading h2 {{
                font-size: 25px;
            }}
            .bold {{
                font-weight: bold;
            }}
            .marksheet table {{
                background: #eee;
                border-collapse: collapse;
                width: 100%;
                margin-top: 20px;
            }}
            .marksheet table tr td {{
                border: 1px solid #fff;
                font-size: 14px;
                padding: 3px 3px;
            }}
            .marksheet table tr th {{
                background-color: #aeb7be;
                text-align: left;
                padding: 3px 3px;
                border: 1px solid #fff;
            }}
            .marksheet table tr:nth-child(odd) {{
                background-color: #dfe0e2;
            }}

            .footer {{
                margin-top: 4rem;
                text-align: center;
                font-size: 1rem;
            }}
        </style>
        <body>
            <div class="main_div">
                <!-- Header -->
                <div class="header-container">
                    <img
                        class="header-img"
                        src="https://outr.ac.in/public/uploads/logo_4.png"
                        width="100"
                        alt=""
                    />
                    <div class="header_right fl_r">
                        <div class="flag clear flag_left fl_l">
                            <h3 class="green_background">Odisha University of Technology and Research</h3>
                        </div>
                    </div>
                </div>
                <!-- Fast Heading -->
                <div class="fast_heading">
                    <h1>6th Semester Marksheet</h1>
                </div>
                <!-- Student Detials -->
                <div class="student_detials">
                    <table>
                        <tr>
                            <td>Roll No.</td>
                            <td>{marks_dict["student_id"]}</td>
                            <td>Name</td>
                            <td>{marks_dict["name"]}</td>
                        </tr>
                        <tr>
                            <td>Branch</td>
                            <td>CSE</td>
                            <td>CGPA</td>
                            <td class="bold">{marks_dict["cgpa"]}</td>
                        </tr>
                    </table>
                </div>
                <div class="second_heading">
                    <h2>Subject Wise Grade</h2>
                </div>

                <div class="marksheet">
                    <table>
                        <tr>
                            <th>Code</th>
                            <th>Subject</th>
                            <th>Grade</th>
                        </tr>
                        {subjects_string}
                    </table>
                </div>

                <!-- footer -->
                <div class="footer">
                    <p>
                        &copy; Odisha University of Technology and Research. All
                        Rights Reserved.
                    </p>
                </div>
            </div>
        </body>
    </html>

    '''

    return html_string


if __name__ == "__main__":
    # print("prediction is", get_prediction('images/image3.jpeg'))
    # print("barcode is", barcode_reader('images/barcode.png'))
    # make_certificate(101)
    student_ids = [2111100398, 2111100399, 2111100400, 2111100401, 2111100402, 2111100403, 2111100404, 2111100405, 2111100406, 2111100407, 2111100408, 2111100409, 2111100410, 2111100411, 2111100412, 2111100413, 2111100414, 2111100415, 2111100416, 2111100417, 2111100418, 2111100419, 2111100420, 2111100421, 2111100422, 2111100423, 2111100424, 2111100425, 2111100426, 2111100427, 2111100428,
                   2111100429, 2111100430, 2111100431, 2111100432, 2111100433, 2111100434, 2111100435, 2111100436, 2111100437, 2111100438, 2111100439, 2111100440, 2111100441, 2111100442, 2111100443, 2111100444, 2111100445, 2111100446, 2111100447, 2111100448, 2111100449, 2111100450, 2111100451, 2111100452, 2111100453, 2111100454, 2111100455, 2111100456, 2111100457, 2111100458, 2111100459, 2111100460]

    for student_id in student_ids:
        updateCGPA(student_id=student_id)
