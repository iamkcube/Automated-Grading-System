from openpyxl import load_workbook
from PIL import Image, ImageFont, ImageDraw
import pandas as pd
from xhtml2pdf import pisa
from prediction import get_prediction
from barcode_reader import barcode_reader

'''Global Variables'''
FONT_FILE = ImageFont.truetype(r"font/Silentha OT.ttf", 150)
FONT_COLOR = "#FFFFFF"

marks_excel_sheet_filename = r"database/Marks copy.xlsx"
students_excel_sheet_filename = r"database/Students copy.xlsx"
marks_excel_sheet = load_workbook(filename=marks_excel_sheet_filename)
students_excel_sheet = load_workbook(filename=students_excel_sheet_filename)


def update_marks_sheet(student_id, subject, endsem_marks):
    sheet = marks_excel_sheet.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        if row[0].value == student_id and row[1].value == subject:
            row[2].value = endsem_marks
            break

    marks_excel_sheet.save(filename=marks_excel_sheet_filename)


def update_students_sheet(student_id, sem, value):
    sheet = students_excel_sheet.active

    row_num = 0

    if sem == 1:
        row_num = 4
    elif sem == 2:
        row_num = 5
    elif sem == 3:
        row_num = 6
    elif sem == 4:
        row_num = 7
    elif sem == 5:
        row_num = 8
    elif sem == 6:
        row_num = 9
    elif sem == 7:
        row_num = 10
    elif sem == 8:
        row_num = 11
    else:
        row_num = 12

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13):
        if row[0].value == student_id:
            row[row_num].value = value
            break

    students_excel_sheet.save(filename=students_excel_sheet_filename)


def getCredits(sem=0):
    if sem == 1:
        return 20.5
    elif sem == 2:
        return 20.5
    elif sem == 3:
        return 23
    elif sem == 4:
        return 20.5
    elif sem == 5:
        return 22.5
    elif sem == 6:
        return 20
    elif sem == 7:
        return 19
    elif sem == 8:
        return 17


def updateCGPA(student_id):
    sheet = students_excel_sheet.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13, values_only=True):
        if row[0] == student_id:
            total_cgpa = 0
            total_credits = 0
            # sum([(row[i]*getCredits(i-3))/getCredits() for i in range(4, 12) if row[i] is not None])

            for i in range(4, 12):
                if row[i] is None:
                    break
                credits = getCredits(i-3)
                total_cgpa += row[i] * credits
                total_credits += credits

            total_cgpa = total_cgpa/total_credits if total_credits != 0 else 0

            print(total_cgpa)

    update_students_sheet(student_id=student_id, sem=0,
                          value=f"{total_cgpa:.2f}")


def make_certificate(student_id):
    students_sheet = students_excel_sheet.active
    marks_sheet = marks_excel_sheet.active

    marks_dict = {}

    students_df = pd.read_excel(students_excel_sheet_filename)
    marks_df = pd.read_excel(marks_excel_sheet_filename)

    filtered_data = students_df[students_df['Student ID'] == student_id]
    if not filtered_data.empty:
        name = filtered_data.iloc[0]['Student Name']
        cgpa = filtered_data.iloc[0]['CGPA']

    student_101_data = marks_df[marks_df['Student ID'] == student_id]
    grades_dict = student_101_data.set_index(
        'Subject Name')['Grades'].to_dict()

    marks_dict["student_id"] = student_id
    marks_dict["name"] = name
    marks_dict["cgpa"] = cgpa
    marks_dict["grades"] = grades_dict

    convert_html_to_pdf(marks_dict=marks_dict,
                        pdf_path=fr"certificates/Certificate - {name}.pdf")


def convert_html_to_pdf(marks_dict, pdf_path):
    with open(pdf_path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(
            make_html_string(marks_dict), dest=pdf_file)

    return not pisa_status.err


def make_html_string(marks_dict):
    subjects_string = ""

    for index, (subject, grade) in enumerate(marks_dict["grades"].items(), 1):
        subjects_string += f'''
                    <tr>
                        <td>{index}</td>
                        <td>{subject.upper()}</td>
                        <td>{grade}</td>
                    </tr>'''

    html_string = f'''
    <!DOCTYPE html>
    <html lang="en">
        <style>
            * {{
                margin: 0;
                padding: 0;
            }}
            body {{
                margin: auto;
                margin-block: 2rem;
            }}
            .header-container {{
                margin: 0;
                padding: 0;
            }}
            .header-img{{
                margin: auto;
            }}
            .flag {{
                display: grid;
                background: #007a16;
                height: 41px;
                overflow: hidden;
                border-bottom: 2px solid #70d06e;
            }}
            .flag_left h2 {{
                color: #70d06e;
                padding: 7px 9px;
                font-size: 20px;
            }}
            .green_background {{
                background: #007a16;
                padding: 4px 9px;
                font-size: 22px;
                color: #fff;
            }}
            .official {{
                text-align: right;
                background: #70d06e;
                padding: 5px;
                text-decoration: none;
                color: #fff;
                font-size: 14px;
            }}
            .fast_heading {{
                text-align: center;
                padding: 20px 0;
                padding-bottom: 2px;
                border-bottom: 2px solid #eee;
                font-size: 2rem;
                margin-top: 12px;
            }}

            .student_detials table {{
                background: #eee;
                border-collapse: collapse;
                width: 100%;
                margin-top: 40px;
            }}
            .student_detials table tr td {{
                border: 1px solid #fff;
                font-size: 14px;
                padding: 3px 3px;
            }}
            .second_heading {{
                text-align: center;
                margin-top: 20px;
            }}
            .second_heading h2 {{
                font-size: 25px;
            }}
            .bold {{
                font-weight: bold;
            }}
            .marksheet table {{
                background: #eee;
                border-collapse: collapse;
                width: 100%;
                margin-top: 20px;
            }}
            .marksheet table tr td {{
                border: 1px solid #fff;
                font-size: 14px;
                padding: 3px 3px;
            }}
            .marksheet table tr th {{
                background-color: #aeb7be;
                text-align: left;
                padding: 3px 3px;
                border: 1px solid #fff;
            }}
            .marksheet table tr:nth-child(odd) {{
                background-color: #dfe0e2;
            }}

            .footer {{
                margin-top: 4rem;
                text-align: center;
                font-size: 1rem;
            }}
        </style>
        <body>
            <div class="main_div">
                <!-- Header -->
                <div class="header-container">
                    <img
                        class="header-img"
                        src="https://outr.ac.in/public/uploads/logo_4.png"
                        width="100"
                        alt=""
                    />
                    <div class="header_right fl_r">
                        <div class="flag clear flag_left fl_l">
                            <h3 class="green_background">Odisha University of Technology and Research</h3>
                        </div>
                    </div>
                </div>
                <!-- Fast Heading -->
                <div class="fast_heading">
                    <h1>4th Semester Marksheet</h1>
                </div>
                <!-- Student Detials -->
                <div class="student_detials">
                    <table>
                        <tr>
                            <td>Roll No.</td>
                            <td>{marks_dict["student_id"]}</td>
                            <td>Name</td>
                            <td>{marks_dict["name"]}</td>
                        </tr>
                        <tr>
                            <td>Branch</td>
                            <td>CSE</td>
                            <td>CGPA</td>
                            <td class="bold">{marks_dict["cgpa"]}</td>
                        </tr>
                    </table>
                </div>
                <div class="second_heading">
                    <h2>Subject Wise Grade</h2>
                </div>

                <div class="marksheet">
                    <table>
                        <tr>
                            <th>Code</th>
                            <th>Subject</th>
                            <th>Grade</th>
                        </tr>
                        {subjects_string}
                    </table>
                </div>

                <!-- footer -->
                <div class="footer">
                    <p>
                        &copy; Odisha University of Technology and Research. All
                        Rights Reserved.
                    </p>
                </div>
            </div>
        </body>
    </html>

    '''

    return html_string


if __name__ == "__main__":
    print("prediction is", get_prediction('images/image3.jpeg'))
    print("barcode is", barcode_reader('images/barcode.png'))
    make_certificate(101)
